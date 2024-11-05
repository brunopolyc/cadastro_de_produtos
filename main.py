import openpyxl
from openpyxl import Workbook
import pathlib


arquivo = pathlib.Path("Cadastro Produtos.xlsx")

if arquivo.exists():
    pass
else:
    arquivo=Workbook()
    sheet = arquivo.active
    sheet['A1'] = "NOME PRODUTO"
    sheet['B1'] = "TIPO"
    sheet['C1'] = "UNIDADES POR TIPO"

    arquivo.save("Cadastro Produtos.xlsx")


produto = input("Nome do Produto: ")
tipo = input("Tipo de pacote: ")
unidade = int(input ("Unidades por Tipo: "))

try:
    arquivo = openpyxl.load_workbook("Cadastro Produtos.xlsx")
    sheet = arquivo.active
    sheet.cell(column=1, row=sheet.max_row+1, value=produto)
    sheet.cell(column=2, row=sheet.max_row, value=tipo)
    sheet.cell(column=3, row=sheet.max_row, value=unidade)

    arquivo.save(r"Cadastro Produtos.xlsx")
    print("Dados salvos com sucesso!")
except:
    print("Erro ao Salvar os Dados")