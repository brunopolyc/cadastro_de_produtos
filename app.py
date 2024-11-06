import customtkinter as ctk
import tkinter as tk
from tkinter import *
from tkinter import messagebox
import openpyxl
from openpyxl import Workbook
import pathlib



ctk.set_appearance_mode("System")
ctk.set_default_color_theme("blue")

class App(ctk.CTk):
    def __init__(self):
        super().__init__()
        self.layout_config()
        self.appearence()
        self.all_system()

    def layout_config(self):
        self.title ("Sistema de Cadastro de Produtos")
        self.geometry("500x700")

    def appearence(self):
        self.lb_apm = ctk.CTkLabel(self, text="Tema", bg_color="transparent", text_color=["#000", "#fff"], font=("Neuzeit Grotesk Light", 16)).place(x=300, y=0)
        self.opt_apm = ctk.CTkOptionMenu(self, values=["Light", "Dark", "System"], command = self.change_apm, font=("Neuzeit Grotesk Light", 16)). place(x=350, y=0)

    
    def all_system(self):
        frame = ctk.CTkFrame(self, width=700, height=70, corner_radius=0, bg_color="#000578", fg_color="#000578")
        frame.place(x=0, y=30)
        title = ctk.CTkLabel(frame, text= "Cadastrar de Produtos", font=("BebasNeue-Regular", 28), text_color="#fff").place(x=150, y=24)
        subtitle = ctk.CTkLabel(self, text= "Por favor preencher as informações abaixo:", font=("Neuzeit Grotesk Bold", 12), text_color=["#000","#fff"]).place(x=50, y=100)
        
        #inputs_texto
        self.name_value = ctk.StringVar()
        self.tipo_value = ctk.StringVar()

        # Usando as variáveis StringVar nos CTkEntry com o parâmetro textvariable
        name_input = ctk.CTkEntry(self, width=400, height=50, corner_radius=20,
                                  textvariable=self.name_value, 
                                  font=("Neuzeit Grotesk Light", 16),
                                  fg_color="transparent", border_color="#006eff")
        name_input.pack(pady=20)

        tipo_input = ctk.CTkEntry(self, width=200, height=50, corner_radius=20,
                                  textvariable=self.tipo_value,
                                  font=("Neuzeit Grotesk Light", 16),
                                  fg_color="transparent", border_color="#006eff")
        tipo_input.pack(pady=20)

        #input_select e text

        unidade_input = ctk.CTkComboBox(self, values=["PCT","CX","UNI"], font=("Neuzeit Grotesk Light", 16), width=150, height=50, corner_radius=20, border_color="#006eff", dropdown_fg_color="#006eff", dropdown_hover_color="#000578")
        unidade_input.set("UNI")
        observaçao_input = ctk.CTkTextbox(self, width=400, height=200, font=("BebasNeue-Light", 16), border_color="#006eff", border_width=2, fg_color="transparent", corner_radius=20)
        
        ficheiro = pathlib.Path("Cadastro Produtos.xlsx")

        if ficheiro.exists():
            pass

        else:            
            ficheiro = Workbook()                    
            sheet = ficheiro.active
            sheet.title = "Produtos"

            sheet['A1'] = "NOME PRODUTO"
            sheet['B1'] = "TIPO"
            sheet['C1'] = "UNIDADES POR EMB."
            sheet['D1'] = "OBS."

            ficheiro.save("Cadastro Produtos.xlsx")

        def clear():
            self.name_value.set("")
            self.tipo_value.set("")
            observaçao_input.delete(0.0, END)

        def submit():
            # Pegando dados dos inputs
            name = self.name_value.get().upper()
            tipo = self.tipo_value.get()
            obs = observaçao_input.get(0.0, END)
            unit = unidade_input.get()
            
            # Verifica se os campos obrigatórios estão preenchidos
            if name == "" or tipo == "" or unit == "":
                error_campos = ctk.CTkLabel(self, text="Por favor preencha todos os campos obrigatórios (*)", 
                                            font=("Neuzeit Grotesk Bold", 12), text_color=["#FF6961", "#FF6961"])
                error_campos.place(x=50, y=530)
                return  

            # Carrega o arquivo Excel
            ficheiro = openpyxl.load_workbook("Cadastro Produtos.xlsx")
            folha = ficheiro.active

            # Verifica se o produto já existe na coluna "A"
            produto_existe = False
            for row in folha.iter_rows(min_row=2, max_col=1, values_only=True):  # Assumindo que os nomes começam na linha 2
                if row[0] == name:
                    produto_existe = True
                    break

            if produto_existe:
                messagebox.showerror("Sistema", "Este produto já está cadastrado!")
            else:
                nova_linha = folha.max_row + 1
                folha.cell(column=1, row=nova_linha, value=name)
                folha.cell(column=2, row=nova_linha, value=tipo)
                folha.cell(column=3, row=nova_linha, value=unit)
                folha.cell(column=4, row=nova_linha, value=obs)
                
                ficheiro.save("Cadastro Produtos.xlsx")
                           
                messagebox.showinfo("Sistema", "Produto Cadastrado com Sucesso!")
                self.name_value.set("")
                self.tipo_value.set("")
                observaçao_input.delete(0.0, END)

        #textos variaveis
        name_value = StringVar()
        tipo_value = StringVar()
        

        #btn
        bnt_submit = ctk.CTkButton(self, text="salvar".upper(), font=("Neuzeit Grotesk Bold", 14), command=submit, fg_color="#151", hover_color="#131", width=100)
        bnt_limpar = ctk.CTkButton(self, text="limpar".upper(), font=("Neuzeit Grotesk Bold", 14), command=clear, fg_color="#555", hover_color="#333", width=100)
        
        #campos
        lb_name = ctk.CTkLabel(self, text= "Nome do Produto:*", font=("Neuzeit Grotesk Light", 16), text_color=["#006eff","#fff"])
        lb_unidade = ctk.CTkLabel(self, text= "Unidades por Embalagem:*", font=("Neuzeit Grotesk Light", 16), text_color=["#006eff","#fff"])
        lb_tipo = ctk.CTkLabel(self, text= "Tipo:*", font=("Neuzeit Grotesk Light", 16), text_color=["#006eff","#fff"])
        lb_observacao = ctk.CTkLabel(self, text= "Observações:", font=("Neuzeit Grotesk Light", 16), text_color=["#006eff","#fff"])

        #posições inputs na janelaA
        lb_name.place(x=50, y=120)
        name_input.place(x=50, y=150)

        lb_unidade.place(x=50, y=210)
        tipo_input.place(x=50,y=240)

        lb_tipo.place(x=270, y=210)
        unidade_input.place(x=270,y=240)

        lb_observacao.place(x=50, y=300)
        observaçao_input.place(x=50, y=330)

        bnt_limpar.place(x=240,y=560)
        bnt_submit.place(x=350,y=560)

    def change_apm(self, new_aparence_mode):
        ctk.set_appearance_mode(new_aparence_mode)

if __name__=="__main__":
    app = App()
    app.mainloop()

